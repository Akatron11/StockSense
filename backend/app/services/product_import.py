"""Excel ürün içe aktarma — parse + validate. DB'ye yazma yapmaz, sadece SKU çakışma kontrolü için okur.

PROCESS.md Faz 4 "Excel import modülü" (2026-08-13 brainstorming ile netleşti) — sadece ilk kurulum/
bulk-seed senaryosu, var olan ürünleri güncellemez. Detay: docs/superpowers/specs/2026-08-13-excel-
product-import-design.md
"""

from dataclasses import dataclass
from datetime import date, datetime
from io import BytesIO

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from ..models import Product

EXPECTED_HEADERS = ["name", "sku", "category", "default_price", "cost_price", "best_before_date"]
MAX_ROWS = 2000
MAX_FILE_SIZE_BYTES = 5 * 1024 * 1024


@dataclass
class ParsedRow:
    name: str
    sku: str
    category: str | None
    default_price: float
    cost_price: float | None
    best_before_date: date | None


@dataclass
class ImportRowError:
    row: int | None
    message: str


def _cell_str(value: object) -> str:
    return str(value).strip() if value is not None else ""


def _is_blank_row(raw: tuple) -> bool:
    return all(cell is None or str(cell).strip() == "" for cell in raw)


def _parse_price(value: object, field_label: str, errors: list[str]) -> float | None:
    text = _cell_str(value)
    if not text:
        errors.append(f"{field_label} zorunlu")
        return None
    try:
        return float(text)
    except ValueError:
        errors.append(f"{field_label} geçerli bir sayı değil")
        return None


def _parse_optional_price(value: object, field_label: str, errors: list[str]) -> float | None:
    text = _cell_str(value)
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        errors.append(f"{field_label} geçerli bir sayı değil")
        return None


def _parse_optional_date(value: object, errors: list[str]) -> date | None:
    if value is None or _cell_str(value) == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _cell_str(value)
    try:
        return datetime.strptime(text, "%Y-%m-%d").date()
    except ValueError:
        errors.append("best_before_date geçerli bir tarih değil (YYYY-MM-DD)")
        return None


def parse_and_validate(
    file_bytes: bytes, company_id: int, db: Session
) -> tuple[list[ParsedRow], list[ImportRowError]]:
    try:
        workbook = load_workbook(BytesIO(file_bytes), data_only=True)
    except Exception:
        return [], [ImportRowError(row=None, message="Dosya okunamadı, geçerli bir .xlsx dosyası olduğundan emin olun")]

    sheet = workbook.active
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    normalized_header = [_cell_str(cell) for cell in (header_row or ())]
    if normalized_header[: len(EXPECTED_HEADERS)] != EXPECTED_HEADERS:
        return [], [
            ImportRowError(
                row=None,
                message="Sütun başlıkları template ile uyuşmuyor, lütfen template'i indirip tekrar deneyin",
            )
        ]

    total_data_rows = max(sheet.max_row - 1, 0)
    if total_data_rows > MAX_ROWS:
        return [], [ImportRowError(row=None, message=f"Dosya çok büyük (maks. {MAX_ROWS} satır)")]

    existing_skus = set(db.scalars(select(Product.sku).where(Product.company_id == company_id)))

    rows: list[ParsedRow] = []
    errors: list[ImportRowError] = []
    seen_skus: dict[str, int] = {}

    row_num = 1
    for raw in sheet.iter_rows(min_row=2, values_only=True):
        row_num += 1
        if not raw or _is_blank_row(raw):
            continue

        padded = (list(raw) + [None] * 6)[:6]
        name_val, sku_val, category_val, price_val, cost_val, bbd_val = padded

        row_errors: list[str] = []

        name = _cell_str(name_val)
        if not name:
            row_errors.append("name zorunlu")

        sku = _cell_str(sku_val)
        if not sku:
            row_errors.append("sku zorunlu")

        category = _cell_str(category_val) or None

        default_price = _parse_price(price_val, "default_price", row_errors)
        cost_price = _parse_optional_price(cost_val, "cost_price", row_errors)
        best_before_date = _parse_optional_date(bbd_val, row_errors)

        if sku:
            if sku in existing_skus:
                row_errors.append(f"SKU zaten kayıtlı ({sku})")
            elif sku in seen_skus:
                row_errors.append(f"SKU tekrarlı ({sku}, satır {seen_skus[sku]} ile çakışıyor)")
            else:
                seen_skus[sku] = row_num

        if row_errors:
            for message in row_errors:
                errors.append(ImportRowError(row=row_num, message=f"Satır {row_num}: {message}"))
            continue

        rows.append(
            ParsedRow(
                name=name,
                sku=sku,
                category=category,
                default_price=default_price,  # type: ignore[arg-type]
                cost_price=cost_price,
                best_before_date=best_before_date,
            )
        )

    if errors:
        return [], errors
    return rows, []
